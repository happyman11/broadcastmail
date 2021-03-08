#%%
###import packages

#https://realpython.com/python-send-email/

import streamlit as st
import smtplib,ssl
import re
import time
import openpyxl 
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase 
from email import encoders


#%%


col1, col2= st.beta_columns(2)
with col1:
    st.write("**Profile :** https://www.rstiwari.com")
with col2:
   st.write("**Blog :** https://tiwari11-rst.medium.com/")

# Text/Title
st.title("Bulk Mail")



#Functions

def read_pdf(path):
    
    if path is not None:
     
        List_email = openpyxl.load_workbook(path)
        
        return(List_email)

    
def check(email):  
  
    regex = '^[a-zA-Z0-9.!#$%&*+/=?^_`{|}~-]+@[a-zA-Z0-9-]+(?:\.[a-zA-Z0-9-]+)*$'
    if(re.search(regex,email)):  
        return(True)       
    else:  
        return(False)

def read_mail_contents_Frm_file(path):
    file_opened = open(path, "r")
    mail_contents=file_opened .read()
    return(mail_contents)
    
def send_mail(email,Sender_email,Subject,Password,attachment_mail):
    
    msg = MIMEMultipart()
    msg['From'] = Sender_email
    passd=Password
    msg['Subject'] = Subject
    msg['To'] = email
    s = smtplib.SMTP('smtp.gmail.com', 587) 
    s.starttls() 
    s.login(msg['From'], passd) 
#greeting with email
    message_greeting="Greetings  "+ email+" ,"
#custom_mail_message
    path_mail_custom_contents="./Mail_Content/Mail_content Custom_message.txt"
    Custom_message_body=read_mail_contents_Frm_file(path_mail_custom_contents)
#fixed_content
    path_mail_fixed_contents="./Mail_Content/Mail_contents_fixed_template.txt"
    fixed_message_body=read_mail_contents_Frm_file(path_mail_fixed_contents)
#file name
    filename = "Week-1 AI-Overview.pdf"
#file path 
    attachment = open("./Resources_attachment/Week-1 AI-Overview.pdf", "rb") 
#attach file
    p = MIMEBase('application', 'octet-stream')
    p.set_payload((attachment).read()) 
    encoders.encode_base64(p)
    p.add_header('Content-Disposition', "attachment; filename= %s" % filename)
    msg.attach(p) 

#attach message
    msg.attach(MIMEText(message_greeting, 'plain'))
    msg.attach(MIMEText(Custom_message_body, 'html'))
    msg.attach(MIMEText(fixed_message_body, 'html'))
    s.login(msg['From'], passd) 
    s.sendmail(msg['From'], msg['To'], msg.as_string()) 
    s.quit()  



st.sidebar.header("Credentials")

Sender_email=st.sidebar.text_input("Sender Email(Note:less Secure apps must be open)","Sender Email")
Sender_email=str(Sender_email).strip()



Password=st.sidebar.text_input("Email Password","Email Password")
Password=str(Password).strip()

st.sidebar.header("Mail Information")
Subject=st.sidebar.text_input("Enter Subject of Email","Enter Subjects...")
Subject=(str(Subject)).strip()


#st.sidebar.subheader("Mail Attachment __NOT FUNCTIONAL")
#attachment_mail = st.sidebar.file_uploader("Upload your input file", type=["xlsx","pfd","doc","png","jpeg","py","ipnyb"])

st.sidebar.subheader("Reciever Xlsx(.xlsx format only)")
uploaded_file = st.sidebar.file_uploader("Upload your input file", type=["xlsx"])


if(st.sidebar.button("Sent Bulk Mail to every one")):
   
   if(not(check(Sender_email))):
      st.write(Sender_email," -->Incorrect Email Enter Again")
      Sender_email=""

   emails_list=read_pdf(uploaded_file)
   sheet_obj = emails_list.active 
   m_row = sheet_obj.max_row 
   count=0
   for i in range(1, m_row + 1): 
    cell_obj = sheet_obj.cell(row = i, column = 1)
    email=cell_obj.value
    if(check(email)):
       
       send_mail(email,Sender_email,Subject,Password,attachment_mail)
       st.write(count,email," -->Email Sent")
       st.success("Mail Sent")
       count=count+1
    else:
       st.write(count,email," -->Email Sent Failed")
       st.error("Mail Sent Failed Incorrect Email") 
       count=count+1
         
   st.success("All Mail Sent...")         
mannual=st.sidebar.text_input("Enter_Email","") 
if(st.sidebar.button("Sent Mail")):
  email_mannual=(str(mannual)).strip()
  if(check(email_mannual)):
       
      send_mail(email_mannual,Sender_email,Subject,Password,attachment_mail)
      st.write(email_mannual," -->Email Sent")
      st.success("Mail Sent")
  else:
      st.write(email_mannual," -->Email Sent Failed")
      st.error("Mail Sent Failed Incorrect Email")
    
    
    
